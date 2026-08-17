import json
import re
import zipfile
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

TEMPLATE = r'C:\Users\jsona\OneDrive\Documents\Claude\DailyReportApp\template\PR439-Daily-Work-Report-TEMPLATE.xlsx'
OUT = r'C:\Users\jsona\OneDrive\Documents\Claude\DailyReportApp\print-layout.json'

# ---- Resolve theme colors from the workbook's theme XML ----
with zipfile.ZipFile(TEMPLATE) as z:
    theme_xml = z.read('xl/theme/theme1.xml').decode('utf-8')

srgb = re.findall(r'<a:(\w+)>\s*<a:srgbClr val="([0-9A-Fa-f]{6})"/>', theme_xml)
sysc = re.findall(r'<a:(\w+)>\s*<a:sysClr val="\w+" lastClr="([0-9A-Fa-f]{6})"/>', theme_xml)
raw = dict(srgb + sysc)
# OOXML theme element order -> Excel color-picker theme index order
THEME_ORDER = ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6', 'hlink', 'folHlink']
THEME_HEX = [raw[name] for name in THEME_ORDER]


def apply_tint(hex_color, tint):
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    def adj(c):
        if tint is None or tint == 0:
            return c
        if tint > 0:
            return round(c + (255 - c) * tint)
        return round(c * (1 + tint))
    return '#%02X%02X%02X' % (adj(r), adj(g), adj(b))


def resolve_color(color):
    if color is None:
        return None
    try:
        if color.type == 'rgb' and color.rgb and color.rgb not in ('00000000',):
            hexval = color.rgb[-6:]
            return apply_tint(hexval, getattr(color, 'tint', 0))
        if color.type == 'theme':
            hexval = THEME_HEX[color.theme]
            return apply_tint(hexval, getattr(color, 'tint', 0))
        if color.type == 'indexed':
            return None  # rare in this file, skip
    except Exception:
        return None
    return None


def col_width_to_pt(width):
    if width is None:
        width = 8.43
    px = round(width * 7 + 5)
    return round(px * 0.75, 2)


def border_side(side):
    if side is None or side.style is None:
        return None
    return {'style': side.style, 'color': resolve_color(side.color) or '#000000'}


def resolve_col_width(ws, col_index):
    # column_dimensions entries can cover a min..max RANGE (e.g. D covering
    # D-H as one <col> definition) -- a plain letter lookup misses that for
    # every letter in the range except the dict's own key, so scan ranges.
    for dim in ws.column_dimensions.values():
        if dim.min is not None and dim.max is not None and dim.min <= col_index <= dim.max:
            if dim.width:
                return dim.width
    return None


def extract_sheet(ws):
    max_row = ws.max_row
    max_col = ws.max_column

    columns = []
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        width = resolve_col_width(ws, c)
        columns.append({'col': c, 'letter': letter, 'widthPt': col_width_to_pt(width)})

    default_row_height = ws.sheet_format.defaultRowHeight or 15.0
    rows = []
    for r in range(1, max_row + 1):
        dim = ws.row_dimensions.get(r)
        height = dim.height if dim and dim.height else default_row_height
        rows.append({'row': r, 'heightPt': round(height, 2)})

    merges = [str(m) for m in ws.merged_cells.ranges]
    merge_anchor_set = {str(m).split(':')[0] for m in ws.merged_cells.ranges}

    cells = {}
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            coord = cell.coordinate
            has_value = cell.value is not None
            has_fill = cell.fill and cell.fill.patternType == 'solid'
            b = cell.border
            has_border = any(
                (b.top and b.top.style) or (b.bottom and b.bottom.style) or
                (b.left and b.left.style) or (b.right and b.right.style)
                for _ in [0]
            )
            is_merge_anchor = coord in merge_anchor_set
            if not (has_value or has_fill or has_border or is_merge_anchor):
                continue

            entry = {}
            if has_value:
                entry['text'] = str(cell.value)
            if has_fill:
                fg = resolve_color(cell.fill.fgColor)
                if fg:
                    entry['fill'] = fg
            border = {}
            for side_name in ('top', 'right', 'bottom', 'left'):
                side = getattr(b, side_name)
                resolved = border_side(side)
                if resolved:
                    border[side_name] = resolved
            if border:
                entry['border'] = border

            font = cell.font
            font_entry = {}
            if font.bold:
                font_entry['bold'] = True
            if font.italic:
                font_entry['italic'] = True
            if font.sz and font.sz != 10.0:
                font_entry['size'] = font.sz
            fc = resolve_color(font.color) if font.color else None
            if fc and fc != '#000000':
                font_entry['color'] = fc
            if font_entry:
                entry['font'] = font_entry

            align = cell.alignment
            align_entry = {}
            if align.horizontal:
                align_entry['h'] = align.horizontal
            if align.vertical:
                align_entry['v'] = align.vertical
            if align.wrapText:
                align_entry['wrap'] = True
            # Vertical (rotated) text -- the contractor name headers in C5:H5
            # are textRotation 90, and they have to stay rotated in the PDF or
            # they overflow their narrow columns.
            if align.textRotation:
                align_entry['rot'] = align.textRotation
            if align_entry:
                entry['align'] = align_entry

            if entry:
                cells[coord] = entry

    return {
        'maxRow': max_row,
        'maxCol': max_col,
        'columns': columns,
        'rows': rows,
        'merges': merges,
        'cells': cells,
    }


def page_setup(ws):
    # The print area is what actually lands on paper -- the sheets both have
    # stray formatting well past it (max_col 38 vs. print area ending at Q),
    # and rendering those extra columns is what blew the page proportions out.
    area = ws.print_area
    if isinstance(area, (list, tuple)):
        area = area[0] if area else None
    m = re.search(r'\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)', str(area or ''))
    if m:
        last_col = column_index_from_string(m.group(3))
        last_row = int(m.group(4))
    else:
        last_col, last_row = ws.max_column, ws.max_row

    pm = ws.page_margins
    return {
        'printArea': str(area),
        'lastCol': last_col,
        'lastRow': last_row,
        'orientation': ws.page_setup.orientation or 'portrait',
        'scale': ws.page_setup.scale or 100,
        # openpyxl reports margins in inches; points are what the renderer uses.
        'marginsPt': {
            'left': round((pm.left or 0) * 72, 2),
            'right': round((pm.right or 0) * 72, 2),
            'top': round((pm.top or 0) * 72, 2),
            'bottom': round((pm.bottom or 0) * 72, 2),
        },
    }


wb = openpyxl.load_workbook(TEMPLATE)
ws1 = wb['Daily Work Report']
ws2 = wb['Daily_Photo_Log']

data = {
    'dailyWorkReport': {**extract_sheet(ws1), **page_setup(ws1)},
    'dailyPhotoLog': {**extract_sheet(ws2), **page_setup(ws2)},
}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=None, separators=(',', ':'))

import os
print('wrote', OUT, os.path.getsize(OUT), 'bytes')
print('sheet1 cells:', len(data['dailyWorkReport']['cells']))
print('sheet2 cells:', len(data['dailyPhotoLog']['cells']))
